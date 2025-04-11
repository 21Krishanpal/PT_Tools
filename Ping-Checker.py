import tkinter as tk
from tkinter import messagebox, filedialog
import subprocess
import threading
import csv
import openpyxl

class PingCheckerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("IP Ping Checker")

        self.ip_label = tk.Label(root, text="Enter IP addresses (comma-separated):")
        self.ip_label.pack()

        self.ip_entry = tk.Entry(root, width=60)
        self.ip_entry.pack(pady=5)

        self.import_button = tk.Button(root, text="Import IPs from XLSX", command=self.import_ips)
        self.import_button.pack(pady=5)

        self.ping_button = tk.Button(root, text="Ping", command=self.start_ping)
        self.ping_button.pack(pady=5)

        self.result_text = tk.Text(root, height=15, width=70)
        self.result_text.pack(pady=10)

        self.export_csv_button = tk.Button(root, text="Export as CSV", command=self.export_csv)
        self.export_csv_button.pack(side=tk.LEFT, padx=10, pady=5)

        self.export_xlsx_button = tk.Button(root, text="Export as XLSX", command=self.export_xlsx)
        self.export_xlsx_button.pack(side=tk.RIGHT, padx=10, pady=5)

        self.results = {}

    def import_ips(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            try:
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active
                ips = []
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str) and cell.strip():
                            ips.append(cell.strip())
                self.ip_entry.delete(0, tk.END)
                self.ip_entry.insert(0, ", ".join(ips))
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read file: {e}")

    def start_ping(self):
        threading.Thread(target=self.ping_ips).start()

    def ping_ips(self):
        self.results.clear()
        self.result_text.delete("1.0", tk.END)
        ips = [ip.strip() for ip in self.ip_entry.get().split(",") if ip.strip()]
        for ip in ips:
            try:
                output = subprocess.run(["ping", "-n", "1", ip], capture_output=True, text=True)
                success = "TTL=" in output.stdout or "Reply from" in output.stdout
                status = "Success" if success else "Failed"
            except Exception as e:
                status = f"Error: {str(e)}"
            self.results[ip] = status
            self.result_text.insert(tk.END, f"{ip}: {status}\n")

    def export_csv(self):
        if not self.results:
            messagebox.showwarning("Warning", "No results to export.")
            return
        filepath = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV Files", "*.csv")])
        if filepath:
            with open(filepath, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["IP Address", "Status"])
                for ip, status in self.results.items():
                    writer.writerow([ip, status])
            messagebox.showinfo("Success", f"Results saved to {filepath}")

    def export_xlsx(self):
        if not self.results:
            messagebox.showwarning("Warning", "No results to export.")
            return
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if filepath:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["IP Address", "Status"])
            for ip, status in self.results.items():
                ws.append([ip, status])
            wb.save(filepath)
            messagebox.showinfo("Success", f"Results saved to {filepath}")

if __name__ == "__main__":
    root = tk.Tk()
    app = PingCheckerApp(root)
    root.mainloop()
